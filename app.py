import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from openpyxl import Workbook

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.all_system()


    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["System", "Dark", "Light"], command=self.change_apm).place(x=50, y=460)

    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de Clientes", font=("Century Gothic bold", 24), text_color="#fff").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        file = pathlib.Path("Clientes.xlsx")

        if file.exists():
            pass
        else:
            file = Workbook()
            sheet = file.active
            sheet['A1']="Nome completo"
            sheet['B1']="Contato"
            sheet['C1']="Idade"
            sheet['D1']="Gênero"
            sheet['E1']="Endereço"
            sheet['F1']="Observações"

            file.save("Clientes.xlsx")


        def submit():
            # Getting data from entries
            name = name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            gender = gender_combobox.get()
            address = address_value.get()
            obs = obs_entry.get(0.0, END)

            if (name == "" or contact == "" or age == "" or address == ""):
                messagebox.showerror("Sistema", "ERRO!\nPreencha todos os campos!")
            else:
                file = openpyxl.load_workbook('Clientes.xlsx')
                sheet = file.active
                sheet.cell(column=1, row=sheet.max_row+1, value=name)
                sheet.cell(column=2, row=sheet.max_row, value=contact)
                sheet.cell(column=3, row=sheet.max_row, value=age)
                sheet.cell(column=4, row=sheet.max_row, value=gender)
                sheet.cell(column=5, row=sheet.max_row, value=address)
                sheet.cell(column=6, row=sheet.max_row, value=obs)

                file.save(r"Clientes.xlsx")
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

        def clear():
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            obs_entry.delete(0.0, END)

        # Texts variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()

        # Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200, textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")

        # Combobox
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino"], font=("Century Gothic bold", 14), width=150)
        gender_combobox.set("Masculino")

        obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        
        # Labels
        lb_name = ctk.CTkLabel(self, text="Nome completo", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_submit = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

        # Positioning elements in the window
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)

        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=180, y=260)


    def change_apm(self, new_appearence_mode):
        ctk.set_appearance_mode(new_appearence_mode)


if __name__ == "__main__":
    app = App()
    app.mainloop()
